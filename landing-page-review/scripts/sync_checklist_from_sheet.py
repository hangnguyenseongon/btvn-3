#!/usr/bin/env python3
"""
Dong bo checklist review landing page tu Google Sheet goc (ban chu - master
copy do nguoi dung duy tri) ve 2 file template cua skill:
    templates/landing_page_review_checklist.csv
    templates/landing_page_review_checklist.xlsx

Google Sheet nguon (public, dang o dang export CSV):
    https://docs.google.com/spreadsheets/d/1q3BLNVSDLwtdxUOWGut1ia1zvKyDPd9RkY3_IBh1Pxw/edit?gid=498393758

Chay lai script nay moi khi ban cap nhat checklist tren Google Sheet, de dong
bo ve template cua skill (dung o moi project).

Cach dung:
    python sync_checklist_from_sheet.py
    python sync_checklist_from_sheet.py --sheet-id <id> --gid <gid>   # doi nguon
"""
import argparse
import csv
import io
import os
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_SHEET_ID = "1q3BLNVSDLwtdxUOWGut1ia1zvKyDPd9RkY3_IBh1Pxw"
DEFAULT_GID = "498393758"

HEADERS = [
    "STT", "Giai đoạn", "Hạng mục đánh giá", "Hạng mục con",
    "Chi tiết công việc", "Tỉ lệ đáp ứng", "Mức độ quan trọng", "Tình trạng",
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def fetch_csv_rows(sheet_id, gid):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    with urllib.request.urlopen(url, timeout=30) as resp:
        raw = resp.read().decode("utf-8-sig")
    return list(csv.reader(io.StringIO(raw)))


def clean_rows(raw_rows):
    """Bo dong tieu de/dong trong, forward-fill cot Giai doan & Hang muc danh gia
    (vi ban goc dung merge-cell nen cac dong ke tiep trong cung nhom de trong)."""
    cleaned = []
    last_phase, last_category = "", ""
    started = False
    for row in raw_rows:
        row = (row + [""] * 7)[:7]
        a, b, c, d, e, f, g = [x.strip() if isinstance(x, str) else x for x in row]

        if not started:
            # bo qua dong tieu de ("Review Website - xyz.com") va dong header goc
            if a.lower().startswith("giai đoạn"):
                started = True
            continue

        if not any([a, b, c, d, e, f, g]):
            continue  # dong trong ngan cach giua cac section

        if a:
            last_phase = a
        if b:
            last_category = b

        cleaned.append([last_phase, last_category, c, d, e, f, g])
    return cleaned


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for i, row in enumerate(rows, start=1):
            writer.writerow([i] + row)


def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist review LDP"
    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(rows, start=1):
        ws.append([i] + row)

    widths = [6, 22, 26, 26, 55, 14, 18, 24]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = WRAP

    ws.freeze_panes = "A2"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="Dong bo checklist review LDP tu Google Sheet goc")
    ap.add_argument("--sheet-id", default=DEFAULT_SHEET_ID)
    ap.add_argument("--gid", default=DEFAULT_GID)
    args = ap.parse_args()

    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    templates_dir = os.path.join(here, "templates")
    os.makedirs(templates_dir, exist_ok=True)

    raw_rows = fetch_csv_rows(args.sheet_id, args.gid)
    rows = clean_rows(raw_rows)

    csv_path = os.path.join(templates_dir, "landing_page_review_checklist.csv")
    xlsx_path = os.path.join(templates_dir, "landing_page_review_checklist.xlsx")
    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)
    print(f"Da dong bo {len(rows)} dong checklist vao:\n  {csv_path}\n  {xlsx_path}")


if __name__ == "__main__":
    main()
