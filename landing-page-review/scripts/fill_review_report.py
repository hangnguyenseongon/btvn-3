#!/usr/bin/env python3
"""
Dien ket qua review thuc te vao ban sao cua file checklist template, xuat ra
file Excel bao cao cuoi cung cho tung du an/khach hang.

File JSON ket qua co dang:
{
  "6":  {"ty_le_dap_ung": true,  "tinh_trang": "Đạt - thông tin đầy đủ, rõ ràng"},
  "9":  {"ty_le_dap_ung": false, "tinh_trang": "Chưa đạt - Google Mobile-Friendly Test báo lỗi font chữ quá nhỏ"},
  "10": {"tinh_trang": "Đạt - Mobile 78/100, Desktop 92/100 (PageSpeed Insights)"}
}
Key la cot STT trong file checklist (xem templates/landing_page_review_checklist.csv).
Chi can dien nhung STT co du lieu thuc te; cac dong con lai giu nguyen (de trong
hoac khong ap dung cho loai website dang review - vi du bo qua nhom "Website/LDP
ban hang" neu dang review web dich vu).

Cach dung:
    python fill_review_report.py --results ket_qua.json \
        --out "Review-LDP-<TenKhachHang>.xlsx"
"""
import argparse
import json
import os

from openpyxl import load_workbook

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_TEMPLATE = os.path.join(SKILL_DIR, "templates", "landing_page_review_checklist.xlsx")

COL_STT = 1
COL_TY_LE = 6
COL_TINH_TRANG = 8


def main():
    ap = argparse.ArgumentParser(description="Dien ket qua review vao checklist va xuat bao cao")
    ap.add_argument("--template", default=DEFAULT_TEMPLATE, help="File checklist goc (mac dinh: template cua skill)")
    ap.add_argument("--results", required=True, help="File JSON ket qua review, key = STT")
    ap.add_argument("--out", required=True, help="Duong dan file Excel bao cao xuat ra")
    args = ap.parse_args()

    with open(args.results, encoding="utf-8") as f:
        results = json.load(f)

    wb = load_workbook(args.template)
    ws = wb.active

    applied = 0
    for row in ws.iter_rows(min_row=2):
        stt_cell = row[COL_STT - 1]
        stt = str(stt_cell.value) if stt_cell.value is not None else None
        if stt in results:
            r = results[stt]
            if "ty_le_dap_ung" in r and r["ty_le_dap_ung"] is not None:
                row[COL_TY_LE - 1].value = bool(r["ty_le_dap_ung"])
            if "tinh_trang" in r:
                row[COL_TINH_TRANG - 1].value = r["tinh_trang"]
            applied += 1

    wb.save(args.out)
    print(f"Da dien ket qua cho {applied}/{len(results)} muc, xuat bao cao vao {args.out}")


if __name__ == "__main__":
    main()
