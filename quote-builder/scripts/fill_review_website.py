#!/usr/bin/env python3
"""
Dien ket qua review landing page (tu skill landing-page-review) vao sheet
"4. Review Website" cua file bao gia.

Dau vao la cung 1 file JSON ket qua dung cho
landing-page-review/scripts/fill_review_report.py (key = STT theo
templates/landing_page_review_checklist.csv cua skill do, vi thu tu 46 hang
muc trong sheet nay giong het checklist goc).

Cach dung:
    python fill_review_website.py --workbook "Bao-gia-KhachHangX.xlsx" \
        --results ket_qua_review.json
"""
import argparse
import json

from openpyxl import load_workbook

SHEET_NAME = "4. Review Website"
ROW_OFFSET = 3  # STT 1 nam o row 4 => row = STT + 3
COL_TY_LE = 5
COL_TINH_TRANG = 7


def main():
    ap = argparse.ArgumentParser(description="Dien ket qua review website vao sheet 4. Review Website")
    ap.add_argument("--workbook", required=True, help="File .xlsx bao gia dang xay dung (ban sao cua template)")
    ap.add_argument("--results", required=True, help="File JSON ket qua review, key = STT (giong skill landing-page-review)")
    args = ap.parse_args()

    with open(args.results, encoding="utf-8") as f:
        results = json.load(f)

    wb = load_workbook(args.workbook)
    ws = wb[SHEET_NAME]

    applied = 0
    for stt_str, r in results.items():
        row = int(stt_str) + ROW_OFFSET
        if "ty_le_dap_ung" in r and r["ty_le_dap_ung"] is not None:
            ws.cell(row=row, column=COL_TY_LE, value=bool(r["ty_le_dap_ung"]))
        if "tinh_trang" in r:
            ws.cell(row=row, column=COL_TINH_TRANG, value=r["tinh_trang"])
        applied += 1

    wb.save(args.workbook)
    print(f"Da dien {applied} muc ket qua review vao sheet '{SHEET_NAME}' trong {args.workbook}")


if __name__ == "__main__":
    main()
