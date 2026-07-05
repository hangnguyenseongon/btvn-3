#!/usr/bin/env python3
"""
Dien bo tu khoa (da gan chu de tu skill keyword-research-search-ads) vao sheet
"2. Target | Search" cua file bao gia, theo dung layout cua template: moi chu
de la 1 khoi 3 cot (Ten chu de | Keyword | Search Volume | [cot trong ngan cach]).
Template co san 6 khoi; neu co nhieu hon 6 chu de, script tu dong them khoi moi
theo cung mau (moi khoi 3 cot).

Dau vao: CSV cung dinh dang voi tagged_keywords.csv cua skill
keyword-research-search-ads (cot: keyword,volume,topic).

Cach dung:
    python fill_target_search.py --workbook "Bao-gia-KhachHangX.xlsx" \
        --keywords tagged_keywords.csv
(script sua truc tiep tren file --workbook, nen chay tren 1 ban sao cua
template, khong sua thang vao templates/quote_template.xlsx)
"""
import argparse
import csv
from collections import defaultdict, OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SHEET_NAME = "2. Target | Search"
BLOCK_WIDTH = 3  # moi khoi chu de: cot Keyword, cot Search Volume, 1 cot trong ngan cach
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_tagged_keywords(path):
    topics = OrderedDict()
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            kw = (row.get("keyword") or "").strip()
            if not kw:
                continue
            topic = (row.get("topic") or "Chưa phân loại").strip() or "Chưa phân loại"
            vol = int(float(row.get("volume") or 0))
            topics.setdefault(topic, []).append((kw, vol))
    # sap xep chu de theo tong volume giam dan, tu khoa trong moi chu de theo volume giam dan
    ordered = sorted(topics.items(), key=lambda kv: sum(v for _, v in kv[1]), reverse=True)
    return [(topic, sorted(rows, key=lambda r: r[1], reverse=True)) for topic, rows in ordered]


def main():
    ap = argparse.ArgumentParser(description="Dien bo tu khoa vao sheet 2. Target | Search")
    ap.add_argument("--workbook", required=True, help="File .xlsx bao gia dang xay dung (ban sao cua template)")
    ap.add_argument("--keywords", required=True, help="File CSV keyword,volume,topic (tu skill keyword-research-search-ads)")
    args = ap.parse_args()

    topics = load_tagged_keywords(args.keywords)
    if not topics:
        raise SystemExit("Khong co du lieu tu khoa hop le trong file --keywords")

    wb = load_workbook(args.workbook)
    ws = wb[SHEET_NAME]

    for i, (topic, rows) in enumerate(topics):
        base_col = 1 + i * BLOCK_WIDTH  # cot bat dau cua khoi (A=1, D=4, G=7, ...)
        kw_col = base_col
        vol_col = base_col + 1

        title_cell = ws.cell(row=1, column=kw_col, value=topic)
        title_cell.font = HEADER_FONT
        title_cell.fill = HEADER_FILL

        ws.cell(row=2, column=kw_col, value="Keyword").font = Font(bold=True)
        ws.cell(row=2, column=vol_col, value="Search Volume").font = Font(bold=True)

        for r, (kw, vol) in enumerate(rows, start=3):
            ws.cell(row=r, column=kw_col, value=kw)
            ws.cell(row=r, column=vol_col, value=vol)

        ws.column_dimensions[get_column_letter(kw_col)].width = 40
        ws.column_dimensions[get_column_letter(vol_col)].width = 14

    wb.save(args.workbook)
    total_kw = sum(len(rows) for _, rows in topics)
    print(f"Da dien {total_kw} tu khoa / {len(topics)} chu de vao sheet '{SHEET_NAME}' trong {args.workbook}")


if __name__ == "__main__":
    main()
