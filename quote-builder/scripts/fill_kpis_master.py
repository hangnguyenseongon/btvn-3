#!/usr/bin/env python3
"""
Dien va tinh toan sheet "1. KPIs | Master" - day chinh la phan "bao gia" /
ke hoach ngan sach - du kien cua chien dich.

Cong thuc (theo dung dinh nghia): voi moi hinh thuc quang cao & moi thang,
nguoi dung nhap: Impr (uoc tinh hien thi), CPC (gia moi click uoc tinh),
CR (ty le chuyen doi uoc tinh), Budget (ngan sach), va tuy chon AOV (gia tri
don hang trung binh). Script tu tinh:
    Click     = Budget / CPC
    CTR       = Click / Impr
    Conv      = Click * CR
    Cost/Conv = Budget / Conv
    Revenue   = Conv * AOV        (neu co AOV)
    ROAS      = Revenue / Budget  (neu co AOV)

Sau khi dien tung hinh thuc, script tu dong:
    - Tinh cot "Total" (ca nam) cho tung hinh thuc: cong don Impr/Click/Budget/
      Conv/Revenue qua cac thang co du lieu, roi tinh lai CTR/CPC/CR/Cost-Conv/
      ROAS tu cac so da cong don (khong lay trung binh cong cua ty le).
    - Tinh nhom "Total" (tong tat ca hinh thuc) cho tung thang va ca nam, theo
      cung nguyen tac cong don roi tinh lai ty le.

Dau vao JSON:
{
  "SEARCH": {
    "Jan": {"impr": 50000, "cpc": 8000, "cr": 0.05, "budget": 20000000},
    "Feb": {"impr": 52000, "cpc": 8200, "cr": 0.05, "budget": 20000000, "aov": 500000}
  },
  "PERFORMANCE MAX": { "Jan": {...} }
}
Ten hinh thuc phai la mot trong: SEARCH, PERFORMANCE MAX, DEMAND GEN, DISPLAY,
DISPLAY - REMARKETING, VIDEO, APP  (khop voi cac nhom co san trong template;
KHONG dua "Total" vao input - dong Total do script tu tinh).

Cach dung:
    python fill_kpis_master.py --workbook "Bao-gia-KhachHangX.xlsx" --plan ke_hoach_kpi.json
"""
import argparse
import json

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET_NAME = "1. KPIs | Master"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_COL = {m: 3 + i for i, m in enumerate(MONTHS)}  # Jan=col C(3) .. Dec=col N(14)
TOTAL_COL = 15  # col O

METRICS = ["Impr", "CTR", "Click", "CPC", "CR", "Conv", "Cost/Conv", "Budget", "AOV", "Revenue", "ROAS"]
METRIC_ROW_OFFSET = {m: i for i, m in enumerate(METRICS)}

GROUP_START_ROW = {
    "Total": 2,
    "SEARCH": 13,
    "PERFORMANCE MAX": 24,
    "DEMAND GEN": 35,
    "DISPLAY": 46,
    "DISPLAY - REMARKETING": 57,
    "VIDEO": 68,
    "APP": 79,
}

FORMATS = [g for g in GROUP_START_ROW if g != "Total"]


def row_of(group, metric):
    return GROUP_START_ROW[group] + METRIC_ROW_OFFSET[metric]


def compute_month(inputs):
    impr = float(inputs["impr"])
    cpc = float(inputs["cpc"])
    cr = float(inputs["cr"])
    budget = float(inputs["budget"])
    aov = inputs.get("aov")

    click = budget / cpc if cpc else 0
    ctr = click / impr if impr else 0
    conv = click * cr
    cost_per_conv = budget / conv if conv else 0

    out = {"Impr": impr, "CTR": ctr, "Click": click, "CPC": cpc, "CR": cr,
           "Conv": conv, "Cost/Conv": cost_per_conv, "Budget": budget}
    if aov:
        aov = float(aov)
        revenue = conv * aov
        roas = revenue / budget if budget else 0
        out.update({"AOV": aov, "Revenue": revenue, "ROAS": roas})
    return out


def aggregate(monthly_values_list):
    """Cong don Impr/Click/Budget/Conv/Revenue, tinh lai cac ty le tu tong."""
    agg = {"Impr": 0.0, "Click": 0.0, "Budget": 0.0, "Conv": 0.0, "Revenue": 0.0}
    has_revenue = False
    for mv in monthly_values_list:
        agg["Impr"] += mv.get("Impr", 0)
        agg["Click"] += mv.get("Click", 0)
        agg["Budget"] += mv.get("Budget", 0)
        agg["Conv"] += mv.get("Conv", 0)
        if "Revenue" in mv:
            agg["Revenue"] += mv["Revenue"]
            has_revenue = True

    result = dict(agg)
    result["CTR"] = agg["Click"] / agg["Impr"] if agg["Impr"] else 0
    result["CPC"] = agg["Budget"] / agg["Click"] if agg["Click"] else 0
    result["CR"] = agg["Conv"] / agg["Click"] if agg["Click"] else 0
    result["Cost/Conv"] = agg["Budget"] / agg["Conv"] if agg["Conv"] else 0
    if has_revenue:
        result["AOV"] = agg["Revenue"] / agg["Conv"] if agg["Conv"] else 0
        result["ROAS"] = agg["Revenue"] / agg["Budget"] if agg["Budget"] else 0
    else:
        del result["Revenue"]
    return result


def write_metric(ws, group, month_or_total, metric, value):
    row = row_of(group, metric)
    col = TOTAL_COL if month_or_total == "Total" else MONTH_COL[month_or_total]
    cell = ws.cell(row=row, column=col, value=round(value, 4))
    if metric in ("CTR", "CR"):
        cell.number_format = "0.00%"
    elif metric == "ROAS":
        cell.number_format = "0.00"
    elif metric in ("Impr", "Click"):
        cell.number_format = "#,##0"
    else:
        cell.number_format = "#,##0"


def main():
    ap = argparse.ArgumentParser(description="Dien va tinh toan sheet 1. KPIs | Master")
    ap.add_argument("--workbook", required=True, help="File .xlsx bao gia dang xay dung (ban sao cua template)")
    ap.add_argument("--plan", required=True, help="File JSON ke hoach KPI theo hinh thuc & thang")
    args = ap.parse_args()

    with open(args.plan, encoding="utf-8") as f:
        plan = json.load(f)

    unknown = set(plan) - set(FORMATS)
    if unknown:
        raise SystemExit(f"Hinh thuc khong hop le: {sorted(unknown)}. Chi chap nhan: {FORMATS}")

    wb = load_workbook(args.workbook)
    ws = wb[SHEET_NAME]

    per_format_monthly = {}   # format -> {month: computed_dict}
    for fmt, months in plan.items():
        computed_by_month = {}
        for month, inputs in months.items():
            if month not in MONTHS:
                raise SystemExit(f"Thang khong hop le: {month}")
            computed = compute_month(inputs)
            computed_by_month[month] = computed
            for metric, value in computed.items():
                write_metric(ws, fmt, month, metric, value)
        # cot Total cua tung hinh thuc
        fmt_total = aggregate(computed_by_month.values())
        for metric, value in fmt_total.items():
            write_metric(ws, fmt, "Total", metric, value)
        per_format_monthly[fmt] = computed_by_month

    # nhom "Total" tong hop tat ca hinh thuc, theo tung thang
    overall_year_rows = []
    for month in MONTHS:
        month_values = [per_format_monthly[fmt][month] for fmt in per_format_monthly if month in per_format_monthly[fmt]]
        if not month_values:
            continue
        month_agg = aggregate(month_values)
        overall_year_rows.append(month_agg)
        for metric, value in month_agg.items():
            write_metric(ws, "Total", month, metric, value)

    if overall_year_rows:
        year_total = aggregate(overall_year_rows)
        for metric, value in year_total.items():
            write_metric(ws, "Total", "Total", metric, value)

    wb.save(args.workbook)
    print(f"Da tinh va dien KPIs cho {len(plan)} hinh thuc vao sheet '{SHEET_NAME}' trong {args.workbook}")


if __name__ == "__main__":
    main()
