#!/usr/bin/env python3
"""
Dung file CSV trung gian (keyword,volume,topic) - da duoc gan chu de -
de xuat ra 1 file Excel hoan chinh:
    - Sheet "Tong hop": toan bo tu khoa, cot Tu khoa / Volume / Chu de
    - Moi chu de rieng mot sheet: cot Tu khoa / Volume, sap xep volume giam dan

Cach dung:
    python build_keyword_report.py --in tagged_keywords.csv --out "Bo-tu-khoa-ThuongHieuX.xlsx"

File CSV dau vao can co header: keyword,volume,topic
"""
import argparse
import csv
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def sanitize_sheet_name(name, existing):
    name = re.sub(r"[\\/*?:\[\]]", "", name).strip() or "Chu de"
    name = name[:31]
    base, i = name, 1
    while name in existing:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    existing.add(name)
    return name


def write_sheet(ws, rows, with_topic_column=False):
    headers = ["Từ khóa", "Volume"] + (["Chủ đề"] if with_topic_column else [])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        row = [r["keyword"], r["volume"]] + ([r["topic"]] if with_topic_column else [])
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.column_dimensions[get_column_letter(1)].width = 45
    ws.column_dimensions[get_column_letter(2)].width = 12
    if with_topic_column:
        ws.column_dimensions[get_column_letter(3)].width = 25
    vol_col = 2
    for row_idx in range(2, len(rows) + 2):
        ws.cell(row=row_idx, column=vol_col).number_format = "#,##0"


def main():
    ap = argparse.ArgumentParser(description="Xuat file Excel bo tu khoa chia theo chu de")
    ap.add_argument("--in", dest="infile", required=True, help="CSV dau vao: keyword,volume,topic")
    ap.add_argument("--out", dest="outfile", required=True, help="Duong dan file .xlsx dau ra")
    args = ap.parse_args()

    rows = []
    with open(args.infile, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            if not r.get("keyword"):
                continue
            rows.append({
                "keyword": r["keyword"].strip(),
                "volume": int(float(r.get("volume") or 0)),
                "topic": (r.get("topic") or "Chưa phân loại").strip() or "Chưa phân loại",
            })

    if not rows:
        sys.exit("File CSV dau vao khong co du lieu hop le.")

    rows.sort(key=lambda r: r["volume"], reverse=True)

    by_topic = defaultdict(list)
    for r in rows:
        by_topic[r["topic"]].append(r)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Tổng hợp"
    write_sheet(summary_ws, rows, with_topic_column=True)

    used_names = {"Tổng hợp"}
    for topic in sorted(by_topic.keys(), key=lambda t: -sum(x["volume"] for x in by_topic[t])):
        topic_rows = sorted(by_topic[topic], key=lambda r: r["volume"], reverse=True)
        ws = wb.create_sheet(sanitize_sheet_name(topic, used_names))
        write_sheet(ws, topic_rows, with_topic_column=False)

    wb.save(args.outfile)
    print(f"Da xuat {len(rows)} tu khoa / {len(by_topic)} chu de vao {args.outfile}")


if __name__ == "__main__":
    main()
